"""
Appends translation glossary data and subtitle issues to 术语表&画面字.xlsx.

Sheet layout (all in sheet '术语表', data starts at row 3):
  A–D  : 人名&称呼        — person names & honorifics
  E–G  : 家族名&产业名&团队&地名 — org / place names
  H–J  : 其他&特殊物品本地化  — other localized terms
  K–N  : 画面字           — on-screen text (filled manually)
  O–S  : 字幕轴问题        — subtitle/timing issues

Each section grows independently; entries are appended below the last
non-empty row in that section's primary column.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import List, Optional

from glossary_extractor import GlossaryData, TermEntry
from post_processor import SubtitleIssue

logger = logging.getLogger(__name__)

SHEET_NAME = "术语表"
DATA_START_ROW = 3   # rows 1 & 2 are headers

# Column indices (1-based, openpyxl convention)
COL = {
    # Person names
    "name_orig":        1,   # A
    "name_trans":       2,   # B
    "name_notes":       3,   # C
    "name_timecode":    4,   # D
    # Org / place names
    "org_orig":         5,   # E
    "org_trans":        6,   # F
    "org_timecode":     7,   # G
    # Other terms
    "other_orig":       8,   # H
    "other_trans":      9,   # I
    "other_timecode":  10,   # J
    # On-screen text (not auto-filled)
    "onscreen_orig":   11,   # K
    "onscreen_trans":  12,   # L
    "onscreen_ep":     13,   # M
    "onscreen_tc":     14,   # N
    # Subtitle issues
    "issue_orig":      15,   # O
    "issue_trans":     16,   # P
    "issue_desc":      17,   # Q
    "issue_ep":        18,   # R
    "issue_tc":        19,   # S
}


def _next_empty_row(ws, primary_col: int) -> int:
    """Return the first row >= DATA_START_ROW where primary_col is empty."""
    row = DATA_START_ROW
    while ws.cell(row=row, column=primary_col).value is not None:
        row += 1
    return row


def append_to_glossary(
    xlsx_path: str | Path,
    glossary: GlossaryData,
    issues: List[SubtitleIssue],
    episode_label: str,
) -> None:
    """
    Open the Excel file, append all entries, and save in-place.
    Existing data and formatting are preserved.
    """
    import openpyxl

    path = Path(xlsx_path)
    if not path.exists():
        logger.error("Excel file not found: %s", path)
        return

    wb = openpyxl.load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        logger.error("Sheet '%s' not found in %s", SHEET_NAME, path)
        return

    ws = wb[SHEET_NAME]

    # --- Person names (A–D) ---
    if glossary.person_names:
        row = _next_empty_row(ws, COL["name_orig"])
        for entry in glossary.person_names:
            ws.cell(row=row, column=COL["name_orig"]).value = entry.original
            ws.cell(row=row, column=COL["name_trans"]).value = entry.translation
            ws.cell(row=row, column=COL["name_notes"]).value = entry.notes or None
            ws.cell(row=row, column=COL["name_timecode"]).value = entry.timecode or None
            row += 1
        logger.info("  Wrote %d person name(s)", len(glossary.person_names))

    # --- Org / place names (E–G) ---
    if glossary.org_place_names:
        row = _next_empty_row(ws, COL["org_orig"])
        for entry in glossary.org_place_names:
            ws.cell(row=row, column=COL["org_orig"]).value = entry.original
            ws.cell(row=row, column=COL["org_trans"]).value = entry.translation
            ws.cell(row=row, column=COL["org_timecode"]).value = entry.timecode or None
            row += 1
        logger.info("  Wrote %d org/place name(s)", len(glossary.org_place_names))

    # --- Other terms (H–J) ---
    if glossary.other_terms:
        row = _next_empty_row(ws, COL["other_orig"])
        for entry in glossary.other_terms:
            ws.cell(row=row, column=COL["other_orig"]).value = entry.original
            ws.cell(row=row, column=COL["other_trans"]).value = entry.translation
            ws.cell(row=row, column=COL["other_timecode"]).value = entry.timecode or None
            row += 1
        logger.info("  Wrote %d other term(s)", len(glossary.other_terms))

    # --- Subtitle issues (O–S) ---
    if issues:
        row = _next_empty_row(ws, COL["issue_orig"])
        for issue in issues:
            ws.cell(row=row, column=COL["issue_orig"]).value = issue.original_text
            ws.cell(row=row, column=COL["issue_trans"]).value = issue.translated_text
            ws.cell(row=row, column=COL["issue_desc"]).value = issue.description
            ws.cell(row=row, column=COL["issue_ep"]).value = episode_label
            ws.cell(row=row, column=COL["issue_tc"]).value = (
                f"{issue.start} --> {issue.end}"
            )
            row += 1
        logger.info("  Wrote %d subtitle issue(s)", len(issues))

    wb.save(path)
    logger.info("Excel updated: %s", path)
